# app/services/excel_report_service.py
import logging
import io
import base64
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from pymongo import MongoClient
from bson import ObjectId
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image
from app.config import MONGO_URI, DATABASE_NAME

logger = logging.getLogger(__name__)

class ExcelReportService:
    def __init__(self):
        try:
            logger.info("Conectando a MongoDB para reportes Excel...")
            self.client = MongoClient(MONGO_URI)
            self.db = self.client[DATABASE_NAME]
            self.receipts = self.db["receipts"]
            self.users = self.db["users"]
            logger.info("Conexion exitosa a la base de datos para reportes")
        except Exception as e:
            logger.error(f"Error al conectar a MongoDB: {e}")
            raise

    def generate_excel_report(self, 
                            start_date: str, 
                            end_date: str, 
                            report_type: str = "general",
                            user_id: Optional[str] = None,
                            user_role: str = "lector",
                            codigo_corresponsal: Optional[str] = None) -> bytes:
        """
        Genera reporte Excel completo
        
        Args:
            start_date: Fecha inicio (YYYY-MM-DD)
            end_date: Fecha fin (YYYY-MM-DD)
            report_type: general, daily, weekly, monthly
            user_id: ID del usuario (para lectores)
            user_role: Rol del usuario (admin, operador, lector)
            codigo_corresponsal: Filtro por corresponsal especifico (admin/operador)
        """
        try:
            logger.info(f"Generando reporte Excel: {report_type} del {start_date} al {end_date}")
            
            # Obtener datos segun permisos
            receipts_data = self._get_receipts_data(start_date, end_date, user_id, user_role, codigo_corresponsal)
            
            if not receipts_data:
                return self._generate_empty_report(start_date, end_date, report_type)
            
            # Crear workbook
            wb = Workbook()
            
            # Crear hojas segun el tipo de reporte
            self._create_summary_sheet(wb, receipts_data, start_date, end_date, report_type, user_role)
            self._create_details_sheet(wb, receipts_data, user_role)
            self._create_analysis_sheet(wb, receipts_data, start_date, end_date, user_role)
            
            if report_type in ["weekly", "monthly"]:
                self._create_temporal_analysis_sheet(wb, receipts_data, report_type)
            
            if user_role in ["admin", "operador"]:
                self._create_corresponsal_sheet(wb, receipts_data)
            
            # Aplicar formato general
            self._apply_general_formatting(wb)
            
            # Convertir a bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info(f"Reporte Excel generado exitosamente: {len(receipts_data)} registros")
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel: {e}")
            raise

    def _get_receipts_data(self, start_date: str, end_date: str, 
                          user_id: Optional[str], user_role: str,
                          codigo_corresponsal: Optional[str]) -> List[Dict]:
        """Obtiene datos de comprobantes segun filtros y permisos"""
        try:
            # Convertir fechas
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            
            # Construir query base
            query = {}
            
            # Filtro por fechas (buscar en multiples formatos)
            date_variations = []
            current_date = start_dt
            while current_date <= end_dt:
                date_variations.extend([
                    current_date.strftime("%d/%m/%Y"),
                    current_date.strftime("%d-%m-%Y"),
                    current_date.strftime("%Y-%m-%d")
                ])
                current_date += timedelta(days=1)
            
            query["fecha"] = {"$in": date_variations}
            
            # Aplicar filtros segun rol
            if user_role == "lector" and user_id:
                query["user_id"] = user_id
            elif codigo_corresponsal and user_role in ["admin", "operador"]:
                query["codigo_corresponsal"] = codigo_corresponsal
            
            # Obtener datos
            receipts = list(self.receipts.find(query).sort("created_at", -1))
            
            # Enriquecer datos
            for receipt in receipts:
                receipt["_id"] = str(receipt["_id"])
                if receipt.get("user_id"):
                    receipt["user_id"] = str(receipt["user_id"])
                
                # Calcular dia de la semana
                try:
                    fecha_str = receipt.get("fecha", "")
                    if "/" in fecha_str:
                        fecha_dt = datetime.strptime(fecha_str, "%d/%m/%Y")
                    elif "-" in fecha_str:
                        fecha_dt = datetime.strptime(fecha_str, "%d-%m-%Y")
                    else:
                        fecha_dt = datetime.strptime(fecha_str, "%Y-%m-%d")
                    
                    receipt["dia_semana"] = fecha_dt.strftime("%A")
                    receipt["semana_num"] = fecha_dt.isocalendar()[1]
                    receipt["mes_num"] = fecha_dt.month
                    receipt["ano"] = fecha_dt.year
                    receipt["fecha_ordenable"] = fecha_dt
                except:
                    receipt["dia_semana"] = "Unknown"
                    receipt["semana_num"] = 0
                    receipt["mes_num"] = 0
                    receipt["ano"] = 0
                    receipt["fecha_ordenable"] = datetime.now()
            
            return receipts
            
        except Exception as e:
            logger.error(f"Error obteniendo datos: {e}")
            return []

    def _create_summary_sheet(self, wb: Workbook, data: List[Dict], 
                            start_date: str, end_date: str, 
                            report_type: str, user_role: str):
        """Crea hoja de resumen ejecutivo"""
        ws = wb.active
        ws.title = "Resumen Ejecutivo"
        
        # Titulo principal
        ws["A1"] = "RIOCAJA SMART - REPORTE DE COMPROBANTES"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:F1")
        
        # Informacion del reporte
        ws["A3"] = "INFORMACION DEL REPORTE"
        ws["A3"].font = Font(size=12, bold=True)
        
        ws["A4"] = "Periodo:"
        ws["B4"] = f"{start_date} al {end_date}"
        ws["A5"] = "Tipo de Reporte:"
        ws["B5"] = report_type.title()
        ws["A6"] = "Fecha de Generacion:"
        ws["B6"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws["A7"] = "Total de Registros:"
        ws["B7"] = len(data)
        
        # Estadisticas generales
        total_valor = sum(float(r.get("valor_total", 0)) for r in data)
        tipos_unicos = len(set(r.get("tipo", "") for r in data))
        
        ws["A9"] = "ESTADISTICAS GENERALES"
        ws["A9"].font = Font(size=12, bold=True)
        
        ws["A10"] = "Valor Total:"
        ws["B10"] = f"${total_valor:,.2f}"
        ws["B10"].font = Font(bold=True, color="2E7D32")
        
        ws["A11"] = "Tipos de Transacciones:"
        ws["B11"] = tipos_unicos
        
        ws["A12"] = "Promedio por Transaccion:"
        ws["B12"] = f"${total_valor/len(data):,.2f}" if data else "$0.00"
        
        # Resumen por tipo
        tipo_stats = {}
        for receipt in data:
            tipo = receipt.get("tipo", "Sin Tipo")
            valor = float(receipt.get("valor_total", 0))
            if tipo not in tipo_stats:
                tipo_stats[tipo] = {"count": 0, "total": 0}
            tipo_stats[tipo]["count"] += 1
            tipo_stats[tipo]["total"] += valor
        
        ws["A14"] = "RESUMEN POR TIPO DE TRANSACCION"
        ws["A14"].font = Font(size=12, bold=True)
        
        headers = ["Tipo", "Cantidad", "Valor Total", "Porcentaje"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=15, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        
        row = 16
        for tipo, stats in sorted(tipo_stats.items(), key=lambda x: x[1]["total"], reverse=True):
            ws.cell(row=row, column=1, value=tipo)
            ws.cell(row=row, column=2, value=stats["count"])
            ws.cell(row=row, column=3, value=f"${stats['total']:,.2f}")
            percentage = (stats["total"] / total_valor * 100) if total_valor > 0 else 0
            ws.cell(row=row, column=4, value=f"{percentage:.1f}%")
            row += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _create_details_sheet(self, wb: Workbook, data: List[Dict], user_role: str):
        """Crea hoja con todos los detalles de transacciones"""
        ws = wb.create_sheet("Detalle de Transacciones")
        
        # Headers base
        headers = [
            "Fecha", "Hora", "Tipo", "Nro. Transaccion", 
            "Valor Total", "Dia Semana", "Semana", "Mes"
        ]
        
        # Headers adicionales para admin/operador
        if user_role in ["admin", "operador"]:
            headers.extend([
                "Codigo Corresponsal", "Nombre Corresponsal", 
                "Nombre Local", "Email Usuario"
            ])
        
        # Escribir headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Escribir datos
        for row, receipt in enumerate(data, 2):
            ws.cell(row=row, column=1, value=receipt.get("fecha", ""))
            ws.cell(row=row, column=2, value=receipt.get("hora", ""))
            ws.cell(row=row, column=3, value=receipt.get("tipo", ""))
            ws.cell(row=row, column=4, value=receipt.get("nro_transaccion", ""))
            ws.cell(row=row, column=5, value=float(receipt.get("valor_total", 0)))
            ws.cell(row=row, column=6, value=receipt.get("dia_semana", ""))
            ws.cell(row=row, column=7, value=receipt.get("semana_num", ""))
            ws.cell(row=row, column=8, value=receipt.get("mes_num", ""))
            
            if user_role in ["admin", "operador"]:
                ws.cell(row=row, column=9, value=receipt.get("codigo_corresponsal", ""))
                ws.cell(row=row, column=10, value=receipt.get("nombre_corresponsal", ""))
                ws.cell(row=row, column=11, value=receipt.get("nombre_local", ""))
                ws.cell(row=row, column=12, value=receipt.get("email_usuario", ""))

    def _create_analysis_sheet(self, wb: Workbook, data: List[Dict], 
                             start_date: str, end_date: str, user_role: str):
        """Crea hoja de analisis con graficos"""
        ws = wb.create_sheet("Analisis y Graficos")
        
        # Titulo
        ws["A1"] = "ANALISIS DE TRANSACCIONES"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:D1")

    def _create_temporal_analysis_sheet(self, wb: Workbook, data: List[Dict], report_type: str):
        """Crea analisis temporal para reportes semanales/mensuales"""
        ws = wb.create_sheet(f"Analisis {report_type.title()}")
        
        if report_type == "weekly":
            ws["A1"] = "ANALISIS SEMANAL"
            ws["A1"].font = Font(size=14, bold=True)
        elif report_type == "monthly":
            ws["A1"] = "ANALISIS MENSUAL"
            ws["A1"].font = Font(size=14, bold=True)

    def _create_corresponsal_sheet(self, wb: Workbook, data: List[Dict]):
        """Crea analisis detallado por corresponsal (solo admin/operador)"""
        ws = wb.create_sheet("Analisis por Corresponsal")
        
        ws["A1"] = "ANALISIS DETALLADO POR CORRESPONSAL"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:G1")

    def _apply_general_formatting(self, wb: Workbook):
        """Aplica formato general a todo el workbook"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.freeze_panes = "A2"

    def _generate_empty_report(self, start_date: str, end_date: str, report_type: str) -> bytes:
        """Genera un reporte vacio cuando no hay datos"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sin Datos"
        
        ws["A1"] = "RIOCAJA SMART - REPORTE SIN DATOS"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:D1")
        
        ws["A3"] = f"No se encontraron transacciones para el periodo:"
        ws["A4"] = f"Desde: {start_date}"
        ws["A5"] = f"Hasta: {end_date}"
        ws["A6"] = f"Tipo de reporte: {report_type}"
        ws["A7"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()

    def get_date_range_options(self) -> Dict[str, Dict]:
        """Obtiene opciones predefinidas de rangos de fechas"""
        now = datetime.now()
        
        return {
            "hoy": {
                "start": now.strftime("%Y-%m-%d"),
                "end": now.strftime("%Y-%m-%d"),
                "label": "Hoy"
            },
            "ayer": {
                "start": (now - timedelta(days=1)).strftime("%Y-%m-%d"),
                "end": (now - timedelta(days=1)).strftime("%Y-%m-%d"),
                "label": "Ayer"
            },
            "esta_semana": {
                "start": (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d"),
                "end": now.strftime("%Y-%m-%d"),
                "label": "Esta Semana"
            },
            "ultimos_30_dias": {
                "start": (now - timedelta(days=30)).strftime("%Y-%m-%d"),
                "end": now.strftime("%Y-%m-%d"),
                "label": "Ultimos 30 Dias"
            }
        }

    def validate_date_range(self, start_date: str, end_date: str) -> Tuple[bool, str]:
        """Valida el rango de fechas"""
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            
            if start_dt > end_dt:
                return False, "La fecha de inicio no puede ser mayor que la fecha de fin"
            
            if end_dt > datetime.now():
                return False, "La fecha de fin no puede ser mayor que la fecha actual"
            
            if (end_dt - start_dt).days > 365:
                return False, "El rango de fechas no puede ser mayor a 1 ano"
            
            return True, ""
            
        except ValueError:
            return False, "Formato de fecha invalido. Use YYYY-MM-DD"

    def get_available_corresponsales_for_reports(self) -> List[Dict]:
        """Obtiene lista de corresponsales disponibles para filtros de reportes"""
        try:
            pipeline = [
                {
                    "$match": {
                        "codigo_corresponsal": {"$exists": True, "$ne": None}
                    }
                },
                {
                    "$group": {
                        "_id": "$codigo_corresponsal",
                        "nombre_corresponsal": {"$first": "$nombre_corresponsal"},
                        "nombre_local": {"$first": "$nombre_local"},
                        "total_transacciones": {"$sum": 1},
                        "valor_total": {"$sum": "$valor_total"},
                        "ultima_transaccion": {"$max": "$created_at"}
                    }
                },
                {
                    "$sort": {"total_transacciones": -1}
                }
            ]
            
            result = list(self.receipts.aggregate(pipeline))
            
            corresponsales = []
            for item in result:
                corresponsales.append({
                    "codigo": item["_id"],
                    "nombre": item.get("nombre_corresponsal", "Sin Nombre"),
                    "local": item.get("nombre_local", "Sin Local"),
                    "total_transacciones": item["total_transacciones"],
                    "valor_total": round(item["valor_total"], 2),
                    "ultima_transaccion": item.get("ultima_transaccion")
                })
            
            return corresponsales
            
        except Exception as e:
            logger.error(f"Error obteniendo corresponsales para reportes: {e}")
            return []

    def get_report_statistics(self, start_date: str, end_date: str, 
                            user_id: Optional[str] = None, 
                            user_role: str = "lector") -> Dict:
        """Obtiene estadisticas rapidas del reporte antes de generarlo"""
        try:
            data = self._get_receipts_data(start_date, end_date, user_id, user_role, None)
            
            if not data:
                return {
                    "total_registros": 0,
                    "valor_total": 0,
                    "tipos_unicos": 0,
                    "promedio_transaccion": 0,
                    "fecha_primera": None,
                    "fecha_ultima": None
                }
            
            total_valor = sum(float(r.get("valor_total", 0)) for r in data)
            tipos_unicos = len(set(r.get("tipo", "") for r in data))
            
            fechas = [r.get("fecha_ordenable") for r in data if r.get("fecha_ordenable")]
            fechas.sort()
            
            return {
                "total_registros": len(data),
                "valor_total": round(total_valor, 2),
                "tipos_unicos": tipos_unicos,
                "promedio_transaccion": round(total_valor / len(data), 2) if data else 0,
                "fecha_primera": fechas[0].strftime("%d/%m/%Y") if fechas else None,
                "fecha_ultima": fechas[-1].strftime("%d/%m/%Y") if fechas else None
            }
            
        except Exception as e:
            logger.error(f"Error obteniendo estadisticas: {e}")
            return {
                "total_registros": 0,
                "valor_total": 0,
                "tipos_unicos": 0,
                "promedio_transaccion": 0,
                "fecha_primera": None,
                "fecha_ultima": None
            }